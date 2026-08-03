# -*- coding: utf-8 -*-
"""Build the frozen SEARCH-36 and SEARCH-48 split workbooks.

The two designs are exact extensions of SEARCH-24.  Workbook authoring is
delegated to the bundled artifact-tool helper so the original table styling is
retained without modifying the source workbook.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_VERSION = "nested-search-splits-1.0"
COLUMNS = (
    "train_start",
    "train_end",
    "test_start",
    "test_end",
    "origin_id",
    "design_set",
    "scheme",
    "stratum",
    "cluster_id",
    "event_label",
    "fit_gap_hours",
    "history_policy",
    "pruning_order",
    "optuna_use",
    "origin_layer",
)
TRAIN_START = pd.Timestamp("2021-01-02 00:00")

EXPANSION_A = (
    (
        "2021-02-24",
        "regular",
        "regular_2021_02_24",
        "Representative late-winter weekday near the February operating center",
    ),
    (
        "2021-03-08",
        "calendar",
        "womens_day_2021",
        "International Women's Day",
    ),
    (
        "2021-03-12",
        "stress",
        "march_wind_2021",
        "Highest daily wind-speed condition in the February-September development interval",
    ),
    (
        "2021-04-19",
        "regular",
        "regular_2021_04_19",
        "Representative spring Monday",
    ),
    (
        "2021-06-21",
        "calendar",
        "trinity_2021",
        "Trinity/observed-holiday condition",
    ),
    (
        "2021-08-04",
        "stress",
        "august_q_peak_2021",
        "Highest daily Q peak and high reactive-power variability",
    ),
    (
        "2021-06-17",
        "regular",
        "regular_2021_06_17",
        "Representative early-summer weekday",
    ),
    (
        "2021-06-28",
        "calendar",
        "constitution_day_2021",
        "Constitution Day",
    ),
    (
        "2021-09-30",
        "stress",
        "late_september_variability_2021",
        "High late-season P-Q intraday variability",
    ),
    (
        "2021-04-22",
        "regular",
        "regular_2021_04_22",
        "Representative late-April weekday",
    ),
    (
        "2021-06-30",
        "regular",
        "regular_2021_06_30",
        "Representative late-June weekday",
    ),
    (
        "2021-09-07",
        "regular",
        "regular_2021_09_07",
        "Representative early-autumn weekday",
    ),
)

EXPANSION_B = (
    (
        "2021-02-22",
        "regular",
        "regular_2021_02_22",
        "Ordinary late-winter Monday",
    ),
    (
        "2021-03-09",
        "calendar",
        "womens_day_2021",
        "Return transition after International Women's Day",
    ),
    (
        "2021-05-13",
        "stress",
        "may_rain_2021",
        "Heavy-precipitation onset preceding the existing 14 May stress origin",
    ),
    (
        "2021-04-10",
        "regular",
        "regular_2021_04_10",
        "Ordinary spring weekend",
    ),
    (
        "2021-06-01",
        "calendar",
        "season_boundary_2021_06",
        "Meteorological-summer boundary",
    ),
    (
        "2021-07-09",
        "stress",
        "july_solar_2021",
        "Highest daily irradiance condition in the development interval",
    ),
    (
        "2021-05-21",
        "regular",
        "regular_2021_05_21",
        "Representative late-spring day",
    ),
    (
        "2021-09-01",
        "calendar",
        "season_boundary_2021_09",
        "Meteorological-autumn boundary",
    ),
    (
        "2021-08-09",
        "stress",
        "august_transition_2021",
        "Large day-to-day combined P-Q change and high intraday variability",
    ),
    (
        "2021-04-21",
        "regular",
        "regular_2021_04_21",
        "Representative spring weekday",
    ),
    (
        "2021-07-05",
        "regular",
        "regular_2021_07_05",
        "Ordinary summer Monday",
    ),
    (
        "2021-08-19",
        "regular",
        "regular_2021_08_19",
        "Representative late-summer weekday",
    ),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _iso(value: Any) -> str:
    timestamp = pd.Timestamp(value)
    return timestamp.strftime("%Y-%m-%dT%H:%M:%S")


def _read_core(path: Path) -> pd.DataFrame:
    with pd.ExcelFile(path) as workbook:
        if workbook.sheet_names != ["24"]:
            raise ValueError(
                f"{path.name} must contain only worksheet '24'; "
                f"found={workbook.sheet_names}"
            )
        frame = pd.read_excel(workbook, sheet_name="24")
    frame.columns = [str(column).strip().lower() for column in frame.columns]
    missing = [column for column in COLUMNS[:-1] if column not in frame.columns]
    if missing:
        raise ValueError(f"{path.name} is missing columns: {missing}")
    if len(frame) != 24:
        raise ValueError(f"{path.name} must contain exactly 24 rows")
    for column in ("train_start", "train_end", "test_start", "test_end"):
        frame[column] = pd.to_datetime(frame[column], errors="raise")
    frame["origin_id"] = pd.to_datetime(
        frame["origin_id"], errors="raise"
    ).dt.strftime("%Y-%m-%d")
    frame["origin_layer"] = "CORE24"
    return frame.loc[:, COLUMNS]


def _expansion_rows(
    definitions: Iterable[tuple[str, str, str, str]],
    *,
    design: str,
    start_order: int,
    layer: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for offset, (date, stratum, cluster_id, label) in enumerate(definitions):
        test_start = pd.Timestamp(f"{date} 00:00")
        rows.append(
            {
                "train_start": TRAIN_START,
                "train_end": test_start - pd.Timedelta(hours=1),
                "test_start": test_start,
                "test_end": test_start + pd.Timedelta(hours=23),
                "origin_id": date,
                "design_set": design,
                "scheme": "search24_expanding",
                "stratum": stratum,
                "cluster_id": cluster_id,
                "event_label": label,
                "fit_gap_hours": 0,
                "history_policy": "through_test_start",
                "pruning_order": start_order + offset,
                "optuna_use": True,
                "origin_layer": layer,
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def _build_frames(core: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    core36 = core.copy()
    core36["design_set"] = "SEARCH-36"
    search36 = pd.concat(
        [
            core36,
            _expansion_rows(
                EXPANSION_A,
                design="SEARCH-36",
                start_order=25,
                layer="EXPANSION_A",
            ),
        ],
        ignore_index=True,
    )

    inherited48 = search36.copy()
    inherited48["design_set"] = "SEARCH-48"
    search48 = pd.concat(
        [
            inherited48,
            _expansion_rows(
                EXPANSION_B,
                design="SEARCH-48",
                start_order=37,
                layer="EXPANSION_B",
            ),
        ],
        ignore_index=True,
    )
    return search36, search48


def _records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in frame.to_dict(orient="records"):
        record = dict(row)
        for column in ("train_start", "train_end", "test_start", "test_end"):
            record[column] = _iso(record[column])
        record["origin_id"] = str(record["origin_id"])
        record["fit_gap_hours"] = int(record["fit_gap_hours"])
        record["pruning_order"] = int(record["pruning_order"])
        record["optuna_use"] = bool(record["optuna_use"])
        records.append(record)
    return records


def _locate_artifact_tool_entry(package_root: Path) -> Path:
    configured = os.environ.get("ARTIFACT_TOOL_ENTRY")
    candidates = [
        Path(configured).expanduser() if configured else None,
        package_root
        / "node_modules"
        / "@oai"
        / "artifact-tool"
        / "dist"
        / "artifact_tool.mjs",
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "node"
        / "node_modules"
        / "@oai"
        / "artifact-tool"
        / "dist"
        / "artifact_tool.mjs",
    ]
    for candidate in candidates:
        if candidate is not None and candidate.exists():
            return candidate.resolve()
    raise FileNotFoundError(
        "Could not locate @oai/artifact-tool. Set ARTIFACT_TOOL_ENTRY to "
        "artifact_tool.mjs before running this builder."
    )


def _locate_node() -> Path:
    configured = os.environ.get("NODE_EXECUTABLE")
    candidates = [
        Path(configured).expanduser() if configured else None,
        Path(shutil.which("node")) if shutil.which("node") else None,
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "node"
        / "bin"
        / "node.exe",
    ]
    for candidate in candidates:
        if candidate is not None and candidate.exists():
            return candidate.resolve()
    raise FileNotFoundError(
        "Could not locate Node.js. Set NODE_EXECUTABLE before running this builder."
    )


def _write_openpyxl_workbook(
    path: Path, frame: pd.DataFrame, *, table_name: str
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        frame.to_excel(writer, sheet_name="24", index=False)
        worksheet = writer.book["24"]
        worksheet.freeze_panes = "A2"
        last_column = get_column_letter(len(frame.columns))
        table = Table(displayName=table_name, ref=f"A1:{last_column}{len(frame) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        for index, column in enumerate(frame.columns, start=1):
            width = max(len(str(column)), *(len(str(value)) for value in frame[column]))
            worksheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 32)


def build_nested_splits(
    package_root: Path, *, backend: str = "openpyxl"
) -> dict[str, Any]:
    package_root = package_root.resolve()
    input_dir = package_root / "Input"
    manifest_dir = package_root / "Results" / "reproducibility"
    manifest_dir.mkdir(parents=True, exist_ok=True)

    source = input_dir / "splits_search24.xlsx"
    if not source.exists():
        raise FileNotFoundError(source)
    core = _read_core(source)
    search36, search48 = _build_frames(core)
    output36 = input_dir / "splits_search36.xlsx"
    output48 = input_dir / "splits_search48.xlsx"
    specification_path = manifest_dir / "nested_split_spec.json"
    specification = {
        "script_version": SCRIPT_VERSION,
        "source_path": str(source.resolve()),
        "source_sha256": _sha256(source),
        "columns": list(COLUMNS),
        "outputs": [
            {
                "design": "SEARCH-36",
                "path": str(output36.resolve()),
                "table_name": "Search36Splits",
                "rows": _records(search36),
            },
            {
                "design": "SEARCH-48",
                "path": str(output48.resolve()),
                "table_name": "Search48Splits",
                "rows": _records(search48),
            },
        ],
    }
    specification_path.write_text(
        json.dumps(specification, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )

    if backend == "artifact-tool":
        helper = Path(__file__).with_name("build_nested_search_workbooks.mjs")
        subprocess.run(
            [
                str(_locate_node()),
                str(helper),
                str(source),
                str(specification_path),
                str(_locate_artifact_tool_entry(package_root)),
            ],
            check=True,
            cwd=package_root,
        )
    elif backend == "openpyxl":
        _write_openpyxl_workbook(output36, search36, table_name="Search36Splits")
        _write_openpyxl_workbook(output48, search48, table_name="Search48Splits")
    else:
        raise ValueError("backend must be 'openpyxl' or 'artifact-tool'")

    for path, expected_rows in ((output36, 36), (output48, 48)):
        with pd.ExcelFile(path) as workbook:
            if workbook.sheet_names != ["24"]:
                raise ValueError(f"{path.name} has unexpected sheets")
            observed = pd.read_excel(workbook, sheet_name="24")
        if len(observed) != expected_rows or list(observed.columns) != list(COLUMNS):
            raise ValueError(f"{path.name} failed post-export verification")

    result = {
        "script_version": SCRIPT_VERSION,
        "source_sha256": _sha256(source),
        "outputs": {
            "SEARCH-36": {
                "path": str(output36.resolve()),
                "sha256": _sha256(output36),
                "rows": 36,
            },
            "SEARCH-48": {
                "path": str(output48.resolve()),
                "sha256": _sha256(output48),
                "rows": 48,
            },
        },
    }
    build_manifest = manifest_dir / "nested_split_build.json"
    build_manifest.write_text(
        json.dumps(result, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    print(f"[nested splits built] {build_manifest}")
    return result


def _build_parser() -> argparse.ArgumentParser:
    package_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--package-root", type=Path, default=package_root)
    parser.add_argument(
        "--backend",
        choices=("openpyxl", "artifact-tool"),
        default="openpyxl",
    )
    return parser


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    args = _build_parser().parse_args()
    build_nested_splits(args.package_root, backend=args.backend)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
